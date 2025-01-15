import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# Configuração do driver do navegador
driver = webdriver.Chrome()

try:
    driver.get('https://consultcpf-devaprender.netlify.app/')

    # Carregar a planilha de clientes
    planilha_clientes = openpyxl.load_workbook('/home/jhon/Documentos/PROJETOS/Python/Project-3/dados_clientes.xlsx')
    pagina_clientes = planilha_clientes['Sheet1']

    # Carregar ou criar a planilha de fechamento
    planilha_fechamento_path = '/home/jhon/Documentos/PROJETOS/Python/Project-3/planilha_fechamento.xlsx'

    try:
        # Tenta carregar a planilha de fechamento existente
        planilha_fechamento = openpyxl.load_workbook(planilha_fechamento_path)
        pagina_fechamento = planilha_fechamento['Sheet1']
    except FileNotFoundError:
        # Se a planilha não existir, cria uma nova
        planilha_fechamento = openpyxl.Workbook()
        pagina_fechamento = planilha_fechamento.active
        pagina_fechamento.title = 'Sheet1'
        # Adiciona o cabeçalho na nova planilha
        pagina_fechamento.append(["Nome", "Valor", "CPF", "Vencimento", "Status", "Data do Pagamento", "Método de Pagamento"])
        planilha_fechamento.save(planilha_fechamento_path)
        print("Planilha de fechamento criada com sucesso!")

    # Iterar pelos clientes na planilha de dados
    for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
        nome, valor, cpf, vencimento = linha

        # Preencher o CPF no campo de pesquisa do site
        sleep(5)
        campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
        campo_pesquisa.clear()
        campo_pesquisa.send_keys(cpf)
        sleep(1)

        # Clicar no botão de pesquisa
        botao_pesquisar = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
        botao_pesquisar.click()
        sleep(4)

        # Capturar o status do cliente
        status_element = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
        status = status_element.text.strip()

        if status.lower() == 'em dia':
            # Capturar data do pagamento e método de pagamento
            data_pagamento_element = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
            metodo_pagamento_element = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")

            # Limpar os dados capturados
            data_pagamento = data_pagamento_element.text.split(":")[1].strip()
            metodo_pagamento = metodo_pagamento_element.text.split(":")[1].strip()

            # Adicionar os dados à planilha de fechamento
            pagina_fechamento.append([nome, valor, cpf, vencimento, 'Em dia', data_pagamento, metodo_pagamento])
        elif status.lower() == 'atrasado':
            # Adicionar os dados como "Pendente" à planilha de fechamento
            pagina_fechamento.append([nome, valor, cpf, vencimento, 'Pendente', None, None])

        # Salvar a planilha de fechamento após cada interação
        planilha_fechamento.save(planilha_fechamento_path)

except Exception as e:
    print(f"Erro durante o processo: {e}")
finally:
    # Encerrar o driver do navegador independentemente de erros
    driver.quit()

print("Processo concluído! Dados salvos na planilha de fechamento.")
